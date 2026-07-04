"""
SEC COMPANIES — FISCAL YEAR END PROPORTION ANALYSIS (FAST VERSION)
====================================================================
Uses ThreadPoolExecutor to fetch multiple companies in parallel.
10 workers = ~10x faster than sequential.

HOW TO USE:
  python SEC_FISCAL_YEAR_ANALYSIS.py

OUTPUT:
  - Printed proportion table on screen
  - fiscal_year_analysis.xlsx (2 sheets: details + summary)

NEEDS: pip install requests pandas openpyxl
"""

import time
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# CONFIGURATION
# ============================================================

TOP_N      = 1000    # ← change to 1000 for full run
MAX_WORKERS = 10    # ← parallel workers (10 = safe for SEC)
USER_AGENT = "Adqvest appservices@adqvest.com"
OUTPUT     = "fiscal_year_analysis.xlsx"

MONTH_NAMES = {
    1:"January", 2:"February", 3:"March",   4:"April",
    5:"May",     6:"June",     7:"July",     8:"August",
    9:"September",10:"October",11:"November",12:"December"
}

# ============================================================
# FETCH FUNCTIONS
# ============================================================

def get_all_companies():
    """Fetch full company list from SEC (ordered by market cap)."""
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    print("📡 Fetching company list from SEC...")
    resp = requests.get(
        "https://www.sec.gov/files/company_tickers.json",
        headers=headers, timeout=30)
    resp.raise_for_status()
    df = pd.DataFrame(resp.json()).T
    print(f"   ✅ {len(df):,} total companies on SEC")
    return df


def fetch_one_company(args):
    """
    Fetch fiscalYearEnd for ONE company.
    Called in parallel by ThreadPoolExecutor.
    Returns a dict with all fields.
    """
    idx, cik, ticker, title = args
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    url     = f"https://data.sec.gov/submissions/CIK{int(cik):010d}.json"

    try:
        resp = requests.get(url, headers=headers,
                            timeout=15, allow_redirects=True)
        if resp.status_code == 200:
            data = resp.json()
            fye  = data.get("fiscalYearEnd", "1231") or "1231"
            name = data.get("name", "") or title
        else:
            fye, name = "1231", title
    except Exception:
        fye, name = "1231", title

    # convert to display
    try:
        month_num  = int(fye[:2])
        day        = int(fye[2:])
        display    = f"{MONTH_NAMES[month_num][:3]} {day}"
        month_name = MONTH_NAMES[month_num]
    except Exception:
        month_num, day, display, month_name = 12, 31, "Dec 31", "December"

    return {
        "Rank":                  idx + 1,
        "CIK":                   int(cik),
        "Ticker":                ticker,
        "Company_Name":          name,
        "FiscalYearEnd_Raw":     fye,
        "FiscalYearEnd_Display": display,
        "FYE_Month":             month_name,
        "FYE_Month_Number":      month_num,
    }


# ============================================================
# SAVE TO EXCEL
# ============================================================

def save_excel(companies_df, summary_df, path):
    wb = Workbook()

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    b_font = Font(name="Arial", size=9)

    # ── Sheet 1: Company Details ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Company Details"
    headers1  = ["Rank","CIK","Ticker","Company_Name",
                 "FiscalYearEnd_Raw","FiscalYearEnd_Display",
                 "FYE_Month","FYE_Month_Number"]
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws1.row_dimensions[1].height = 20

    for ri, (_, row) in enumerate(
            companies_df.sort_values("Rank").iterrows(), 2):
        for ci, col in enumerate(headers1, 1):
            cell = ws1.cell(row=ri, column=ci, value=row[col])
            cell.font = b_font
            cell.alignment = Alignment(vertical="top")

    for ci, w in {1:8,2:12,3:10,4:40,5:16,6:18,7:14,8:18}.items():
        ws1.column_dimensions[ws1.cell(1,ci).column_letter].width = w
    ws1.freeze_panes = "A2"

    # ── Sheet 2: FYE Summary ──────────────────────────────────────
    ws2 = wb.create_sheet("FYE Summary")
    headers2 = ["FYE_Month","FYE_Display","Company_Count",
                "Percentage","Companies_List"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws2.row_dimensions[1].height = 20

    for ri, (_, row) in enumerate(summary_df.iterrows(), 2):
        for ci, col in enumerate(headers2, 1):
            cell = ws2.cell(row=ri, column=ci, value=row[col])
            cell.font = b_font
            cell.alignment = Alignment(
                vertical="top", wrap_text=(ci == 5))

    for ci, w in {1:14,2:14,3:16,4:12,5:80}.items():
        ws2.column_dimensions[ws2.cell(1,ci).column_letter].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


# ============================================================
# MAIN
# ============================================================

def main():
    print("\n" + "="*65)
    print("SEC COMPANIES — FISCAL YEAR END ANALYSIS  (FAST VERSION)")
    print(f"Workers: {MAX_WORKERS} parallel  |  Companies: {TOP_N}")
    print("="*65 + "\n")

    # ── Step 1: Get company list ───────────────────────────────────
    all_companies = get_all_companies()
    top_companies = all_companies.head(TOP_N).reset_index(drop=True)

    # ── Step 2: Fetch all in parallel ─────────────────────────────
    print(f"🔄 Fetching fiscalYearEnd for {TOP_N} companies "
          f"({MAX_WORKERS} at a time)...\n")

    args_list = [
        (idx, int(row["cik_str"]), str(row["ticker"]), str(row["title"]))
        for idx, row in top_companies.iterrows()
    ]

    results   = [None] * TOP_N
    done      = 0
    start     = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_idx = {
            executor.submit(fetch_one_company, args): args[0]
            for args in args_list
        }
        for future in as_completed(future_to_idx):
            idx    = future_to_idx[future]
            result = future.result()
            results[idx] = result
            done += 1

            elapsed = time.time() - start
            rate    = done / elapsed
            eta     = (TOP_N - done) / rate if rate > 0 else 0

            print(f"  [{done:>4}/{TOP_N}] "
                  f"{result['Ticker']:<8} "
                  f"{result['Company_Name'][:30]:<30}  "
                  f"FYE={result['FiscalYearEnd_Raw']} "
                  f"({result['FiscalYearEnd_Display']})  "
                  f"ETA:{eta:.0f}s",
                  flush=True)

    elapsed = time.time() - start
    print(f"\n⏱  Done in {elapsed:.1f}s  "
          f"({elapsed/TOP_N:.2f}s per company)")

    companies_df = pd.DataFrame(results)

    # ── Step 3: Build summary ──────────────────────────────────────
    print("\n" + "="*65)
    print("📊 FISCAL YEAR END DISTRIBUTION")
    print("="*65)

    total       = len(companies_df)
    summary_rows = []
    grouped     = companies_df.groupby(
        ["FYE_Month_Number", "FYE_Month", "FiscalYearEnd_Display"])

    for (month_num, month_name, display), group in sorted(
            grouped, key=lambda x: x[0][0]):
        count        = len(group)
        pct          = count / total * 100
        company_list = ", ".join(
            f"{r['Ticker']}({r['Company_Name'][:15]})"
            for _, r in group.iterrows()
        )
        summary_rows.append({
            "FYE_Month":     month_name,
            "FYE_Display":   display,
            "Company_Count": count,
            "Percentage":    round(pct, 1),
            "Companies_List": company_list,
        })

    # print sorted by count descending
    summary_df = pd.DataFrame(summary_rows).sort_values(
        "Company_Count", ascending=False).reset_index(drop=True)

    print(f"\n  {'FYE Month':<14} {'Display':<10} "
          f"{'Count':>6}  {'%':>6}  Bar")
    print("  " + "-"*55)
    for _, row in summary_df.iterrows():
        bar = "█" * int(row["Percentage"] / 2)
        print(f"  {row['FYE_Month']:<14} {row['FYE_Display']:<10} "
              f"{row['Company_Count']:>6}  {row['Percentage']:>5.1f}%  {bar}")

    print(f"\n  Total analysed : {total} companies")
    print(f"  Most common FYE: {summary_df.iloc[0]['FYE_Month']} "
          f"({summary_df.iloc[0]['FYE_Display']})  "
          f"— {summary_df.iloc[0]['Company_Count']} companies "
          f"({summary_df.iloc[0]['Percentage']}%)")

    # ── Step 4: Save to Excel ─────────────────────────────────────
    print(f"\n💾 Saving to {OUTPUT}...")
    save_excel(companies_df, summary_df, OUTPUT)

    print(f"\n✅ DONE!")
    print(f"   {OUTPUT}")
    print(f"   Sheet 1: Company Details  ({len(companies_df)} rows)")
    print(f"   Sheet 2: FYE Summary      ({len(summary_df)} months)")
    print("\n" + "="*65 + "\n")


if __name__ == "__main__":
    main()