"""
SEC FYE ENRICHMENT — EXCEL VERSION
=====================================
Reads your sample Excel file, fetches fiscalYearEnd from SEC EDGAR API
(parallel, by unique CIK), computes derived columns, saves enriched Excel.

NEW COLUMNS ADDED (6 only):
  - Fiscal_Year_End_Raw    : "0131"
  - Fiscal_Year_Period     : "Feb–Jan"   (fiscal start – fiscal end month)
  - Fiscal_Quarter_Split   : "Q1: Feb–Apr | Q2: May–Jul | ..."
  - Calendar_Year          : "CY2026"
  - Calendar_Quarter       : "Q1" / "Q2" / "Q3" / "Q4"  (calendar)
  - Calendar_Months        : "Jan–Mar" / "Apr–Jun" / "Jul–Sep" / "Oct–Dec"

Calendar Quarter is always:
  Q1 = Jan–Mar  |  Q2 = Apr–Jun  |  Q3 = Jul–Sep  |  Q4 = Oct–Dec

USAGE:
  python sec_fye_excel.py --input your_file.xlsx --output enriched_output.xlsx

NEEDS: pip install requests pandas openpyxl xlrd
"""

import os
import time
import argparse
import requests
import pandas as pd
from datetime import datetime
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# CONFIGURATION
# ============================================================

MAX_WORKERS = 10
USER_AGENT  = "Adqvest appservices@adqvest.com"

MONTH_ABB = {
    1:"Jan", 2:"Feb", 3:"Mar",  4:"Apr",
    5:"May", 6:"Jun", 7:"Jul",  8:"Aug",
    9:"Sep", 10:"Oct",11:"Nov", 12:"Dec"
}

# Calendar quarter definitions — fixed, always Jan-Mar=Q1 etc.
CAL_QUARTER_MONTHS = {
    1: "Jan–Mar",
    2: "Apr–Jun",
    3: "Jul–Sep",
    4: "Oct–Dec",
}

HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT     = Font(name="Arial", size=9)
EXISTING_FILL = PatternFill("solid", start_color="1F4E79")  # navy
NEW_FILL      = PatternFill("solid", start_color="2E7D32")  # dark green

NEW_COLS = [
    "Fiscal_Year_End_Raw",
    "Fiscal_Year_Period",
    "Fiscal_Quarter_Split",
    "Calendar_Year",
    "Calendar_Quarter",
    "Calendar_Months",
]

# ============================================================
# SEC API
# ============================================================

def fetch_fye_for_cik(cik: int) -> str:
    """Returns MMDD e.g. '0131'. Defaults to '1231' on any error."""
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    url = f"https://data.sec.gov/submissions/CIK{cik:010d}.json"
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code == 200:
            fye = resp.json().get("fiscalYearEnd", "1231") or "1231"
            return fye if len(fye) == 4 else "1231"
    except Exception:
        pass
    return "1231"


def fetch_all_fyes(unique_ciks: list) -> dict:
    """Parallel fetch for all unique CIKs. Returns {cik: fye_raw}."""
    results = {}
    total   = len(unique_ciks)
    done    = 0
    start   = time.time()
    print(f"\n🔄 Fetching FYE for {total} unique CIKs ({MAX_WORKERS} workers)...\n")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_fye_for_cik, cik): cik
                      for cik in unique_ciks}
        for future in as_completed(future_map):
            cik     = future_map[future]
            fye     = future.result()
            results[cik] = fye
            done   += 1
            elapsed = time.time() - start
            rate    = done / elapsed if elapsed > 0 else 1
            eta     = (total - done) / rate
            print(f"  [{done:>4}/{total}] CIK={cik:<12} FYE={fye}  ETA:{eta:.0f}s",
                  flush=True)
    return results

# ============================================================
# FYE HELPERS
# ============================================================

def fiscal_year_period(fye_raw: str) -> str:
    """'0131' → 'Feb–Jan'  (fiscal start month – fiscal end month)"""
    try:
        fye_month = int(fye_raw[:2])
        start_m   = (fye_month % 12) + 1
        return f"{MONTH_ABB[start_m]}–{MONTH_ABB[fye_month]}"
    except Exception:
        return "Jan–Dec"


def fiscal_quarter_split(fye_raw: str) -> str:
    """'0131' → 'Q1: Feb–Apr | Q2: May–Jul | Q3: Aug–Oct | Q4: Nov–Jan'"""
    try:
        fye_month = int(fye_raw[:2])
        start_m   = (fye_month % 12) + 1
        parts     = []
        for q in range(1, 5):
            m1 = ((start_m - 1) + (q - 1) * 3) % 12 + 1
            m3 = ((start_m - 1) + (q - 1) * 3 + 2) % 12 + 1
            parts.append(f"Q{q}: {MONTH_ABB[m1]}–{MONTH_ABB[m3]}")
        return " | ".join(parts)
    except Exception:
        return "Q1: Jan–Mar | Q2: Apr–Jun | Q3: Jul–Sep | Q4: Oct–Dec"

# ============================================================
# CALENDAR YEAR / QUARTER LOGIC
# ============================================================
def get_calendar_info(period_of_report: str, fye_raw: str) -> dict:
    """
    Derives Calendar_Year, Calendar_Quarter and Calendar_Months
    using the MAJORITY OVERLAP rule.

    Logic:
    1. Find which fiscal quarter the report belongs to.
    2. Build the 3 months of that fiscal quarter.
    3. Calendar Year = year containing the majority of those 3 months.
       Tie -> later year.
    4. Calendar Quarter = calendar quarter containing the majority of those 3 months.
       Tie -> earlier calendar quarter.
    """

    # ----------------------------------------------------------
    # Parse Period_Of_Report
    # ----------------------------------------------------------
    por = str(period_of_report).strip()

    report_date = None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            report_date = datetime.strptime(por, fmt)
            break
        except ValueError:
            continue

    if report_date is None:
        return {
            "Calendar_Year": "",
            "Calendar_Quarter": "",
            "Calendar_Months": ""
        }

    report_month = report_date.month
    report_year = report_date.year

    # ----------------------------------------------------------
    # Fiscal Year Start Month
    # Example:
    # FYE = Jan -> Fiscal starts Feb
    # ----------------------------------------------------------

    fye_month = int(fye_raw[:2])
    fiscal_start_month = (fye_month % 12) + 1

    # ----------------------------------------------------------
    # Which fiscal quarter does Period_Of_Report belong to?
    # ----------------------------------------------------------

    offset = (report_month - fiscal_start_month) % 12

    fiscal_quarter = offset // 3

    quarter_start_month = ((fiscal_start_month - 1) + fiscal_quarter * 3) % 12 + 1

    # ----------------------------------------------------------
    # Build actual months of that fiscal quarter
    # Example:
    # Feb Apr -> [(2026,2),(2026,3),(2026,4)]
    # ----------------------------------------------------------

    months = []

    year = report_year

    if quarter_start_month > report_month:
        year -= 1

    month = quarter_start_month

    for _ in range(3):

        months.append((year, month))

        month += 1

        if month == 13:
            month = 1
            year += 1

    # ----------------------------------------------------------
    # Calendar Year (Majority Year)
    # ----------------------------------------------------------

    year_counter = Counter(y for y, m in months)

    max_year_count = max(year_counter.values())

    calendar_year = max(
        y for y, c in year_counter.items()
        if c == max_year_count
    )

    # ----------------------------------------------------------
    # Calendar Quarter (Majority Quarter)
    # ----------------------------------------------------------

    quarter_counter = Counter()

    for _, month in months:

        q = (month - 1) // 3 + 1

        quarter_counter[q] += 1

    max_q = max(quarter_counter.values())

    calendar_quarter = min(
        q for q, c in quarter_counter.items()
        if c == max_q
    )

    calendar_months = CAL_QUARTER_MONTHS[calendar_quarter]

    return {
        "Calendar_Year": f"CY{calendar_year}",
        "Calendar_Quarter": f"Q{calendar_quarter}",
        "Calendar_Months": calendar_months,
    }
# ============================================================
# ENRICH DATAFRAME
# ============================================================

def enrich_df(df: pd.DataFrame, cik_col: str, por_col: str) -> pd.DataFrame:
    unique_ciks = df[cik_col].dropna().astype(int).unique().tolist()
    cik_fye_map = fetch_all_fyes(unique_ciks)

    print("\n⚙️  Computing derived columns...")
    rows_data = {col: [] for col in NEW_COLS}

    for _, row in df.iterrows():
        try:
            cik = int(row[cik_col])
        except Exception:
            cik = 0
        fye_raw = cik_fye_map.get(cik, "1231")
        por     = str(row[por_col]).strip()

        cal = get_calendar_info(por, fye_raw)

        rows_data["Fiscal_Year_End_Raw"].append(fye_raw)
        rows_data["Fiscal_Year_Period"].append(fiscal_year_period(fye_raw))
        rows_data["Fiscal_Quarter_Split"].append(fiscal_quarter_split(fye_raw))
        rows_data["Calendar_Year"].append(cal["Calendar_Year"])
        rows_data["Calendar_Quarter"].append(cal["Calendar_Quarter"])
        rows_data["Calendar_Months"].append(cal["Calendar_Months"])

    for col in NEW_COLS:
        df[col] = rows_data[col]

    print(f"   ✅ {len(df):,} rows enriched")
    return df

# ============================================================
# EXCEL SAVE WITH FORMATTING
# ============================================================

COL_WIDTHS = {
    "File_ID": 10, "Cik": 12, "Company_Name": 30,
    "Company_Name_Clean": 25, "Ticker": 10, "Generated_File_Name": 45,
    "Form": 8, "Filed": 14, "Period_Of_Report": 18,
    "Financial_Year": 14, "Quarter": 10, "Items": 12,
    "Item_Description": 35, "Accession": 28, "Primary_Doc": 35,
    "Doc_Type": 10, "Link_File_Type": 14, "Content_Type": 22,
    "Filing_Index_Url": 60, "File_Link": 60,
    "S3_Upload_Status": 18, "S3_Upload_Comments": 20,
    "S3_Upload_Count": 16, "S3_Bucket_Name": 22,
    "S3_Folder_Name": 30, "Chunking_Status": 16,
    "Chunking_Comments": 20, "Runtime": 20,
    # New columns
    "Fiscal_Year_End_Raw": 18,
    "Fiscal_Year_Period": 16,
    "Fiscal_Quarter_Split": 52,
    "Calendar_Year": 14,
    "Calendar_Quarter": 16,
    "Calendar_Months": 16,
}


def save_enriched_excel(df: pd.DataFrame, output_path: str):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    print(f"\n💾 Saving to {output_path}...")
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active

    header_row   = [cell.value for cell in ws[1]]
    new_col_idxs = {i + 1 for i, h in enumerate(header_row) if h in NEW_COLS}

    # Header row formatting
    ws.row_dimensions[1].height = 28
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font      = HEADER_FONT
        cell.fill      = NEW_FILL if col_idx in new_col_idxs else EXISTING_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(cell.value or "", 16)

    # Body formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"   ✅ Saved: {output_path}")

# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Enrich SEC EDGAR Excel with FYE + calendar columns"
    )
    parser.add_argument("--input",   "-i", required=True,
                        help="Path to input Excel file (.xlsx or .xls)")
    parser.add_argument("--output",  "-o", default="sec_fye_enriched.xlsx",
                        help="Output path (default: sec_fye_enriched.xlsx)")
    parser.add_argument("--sheet",   "-s", default=0,
                        help="Sheet name or index (default: first sheet)")
    parser.add_argument("--cik-col", default="Cik",
                        help="CIK column name (default: Cik)")
    parser.add_argument("--por-col", default="Period_Of_Report",
                        help="Period_Of_Report column name (default: Period_Of_Report)")
    args = parser.parse_args()

    print("\n" + "="*65)
    print("SEC FYE ENRICHMENT — EXCEL VERSION")
    print("="*65)

    sheet_arg = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
    print(f"\n📖 Reading: {args.input}  (sheet={sheet_arg})")
    df = pd.read_excel(args.input, sheet_name=sheet_arg, dtype=str)
    print(f"   ✅ {len(df):,} rows  |  {len(df.columns)} columns")
    print(f"   Columns found: {list(df.columns)}")

    for col in [args.cik_col, args.por_col]:
        if col not in df.columns:
            raise ValueError(
                f"Column '{col}' not found. Available: {list(df.columns)}"
            )

    df = enrich_df(df, cik_col=args.cik_col, por_col=args.por_col)

    print("\n" + "="*65)
    print("PREVIEW — NEW COLUMNS")
    print("="*65)
    preview_cols = [args.cik_col, args.por_col] + NEW_COLS
    print(df[[c for c in preview_cols if c in df.columns]].head(10).to_string(index=False))

    save_enriched_excel(df, args.output)

    print("\n" + "="*65)
    print("✅  ENRICHMENT COMPLETE")
    print(f"   Input : {args.input}")
    print(f"   Output: {args.output}")
    print(f"   Rows  : {len(df):,}")
    print(f"   New columns: {', '.join(NEW_COLS)}")
    print("="*65 + "\n")


if __name__ == "__main__":
    main()