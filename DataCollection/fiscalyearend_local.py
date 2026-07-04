import time
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

TOP_N = 10
MAX_WORKERS = 10
USER_AGENT = "Adqvest appservices@adqvest.com"
OUTPUT = r"E:\SEC\DataCollection\sec_filings_output\fiscal_year_analysis.xlsx"

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}

def fiscal_period_label(month_num):
    start_month = month_num % 12 + 1
    return f"{MONTH_NAMES[start_month][:3]}-{MONTH_NAMES[month_num][:3]}"

def get_all_companies():
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    resp = requests.get("https://www.sec.gov/files/company_tickers.json", headers=headers, timeout=30)
    resp.raise_for_status()
    df = pd.DataFrame(resp.json()).T
    return df

def fetch_one_company(args):
    idx, cik, ticker, title = args
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    url = f"https://data.sec.gov/submissions/CIK{int(cik):010d}.json"
    try:
        resp = requests.get(url, headers=headers, timeout=20, allow_redirects=True)
        if resp.status_code == 200:
            data = resp.json()
            fye = data.get("fiscalYearEnd", "1231") or "1231"
            name = data.get("name", "") or title
        else:
            fye = "1231"
            name = title
    except Exception:
        fye = "1231"
        name = title

    try:
        month_num = int(str(fye)[:2])
        day = int(str(fye)[2:])
        display = f"{MONTH_NAMES[month_num][:3]} {day}"
    except Exception:
        month_num, day, display = 12, 31, "Dec 31"

    return {
        "CIK": int(cik),
        "Ticker": ticker,
        "Company_Name": name,
        "SEC_Submission_URL": url,
        "FiscalYearEnd_Raw": fye,
        "FiscalYearEnd_Display": display,
        "FiscalYearPeriod": fiscal_period_label(month_num),
    }

def save_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Details"

    headers = [
        "CIK", "Ticker", "Company_Name", "SEC_Submission_URL",
        "FiscalYearEnd_Raw", "FiscalYearEnd_Display", "FiscalYearPeriod"
    ]

    h_fill = PatternFill("solid", start_color="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    b_font = Font(name="Arial", size=9)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row[col])
            cell.font = b_font
            cell.alignment = Alignment(vertical="top", wrap_text=(col in ["Company_Name", "SEC_Submission_URL"]))

    widths = {1: 12, 2: 12, 3: 35, 4: 48, 5: 16, 6: 18, 7: 16}
    for ci, w in widths.items():
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(path)

def main():
    all_companies = get_all_companies()
    top_companies = all_companies.head(TOP_N).reset_index(drop=True)

    args_list = [
        (idx, int(row["cik_str"]), str(row["ticker"]), str(row["title"]))
        for idx, row in top_companies.iterrows()
    ]

    results = [None] * TOP_N
    start = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_idx = {executor.submit(fetch_one_company, args): args[0] for args in args_list}
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            results[idx] = future.result()

    df = pd.DataFrame(results)
    save_excel(df, OUTPUT)

    elapsed = time.time() - start
    print(df.head(10).to_string(index=False))
    print(f"Saved: {OUTPUT}")
    print(f"Elapsed: {elapsed:.1f}s")

if __name__ == "__main__":
    main()